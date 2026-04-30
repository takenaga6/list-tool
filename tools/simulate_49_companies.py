"""49社シミュレーション再現スクリプト.

evaluate_rank_v2() が、昨日のシミュレーション結果と同じ結果を出すか検証する.

実行方法:
    cd /c/Users/user/projects/list_tool
    python tools/simulate_49_companies.py /c/Users/user/Downloads/2026年営業分析シート__3_.xlsx

要件:
    pip install openpyxl
"""

import sys
import os

# agents/rank_agent.py を import できるようにパス追加
_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(_DIR, ".."))

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl が必要です。以下を実行してください：")
    print("  pip install openpyxl")
    sys.exit(1)

from agents.rank_agent import evaluate_rank_v2


def load_companies(xlsx_path: str) -> list[dict]:
    """エクセルから49社データを読み込む."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["契約企業分析"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    companies = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 1).value
        if not name:
            continue
        company = {}
        for c in range(1, ws.max_column + 1):
            company[headers[c - 1]] = ws.cell(row, c).value
        companies.append(company)
    return companies


def to_company_info(c: dict) -> dict:
    """エクセル行を company_info 形式に変換.

    注意：エクセルから読み込んだ数値はfloat（60.0等）なので、
    rank_agent.py側のint変換で「60.0」→「600」と誤読される問題を防ぐため、
    ここでint変換してから渡す。
    """
    emp = c.get("従業員数")
    # floatをintに変換（例：60.0 → 60）。Noneや非数値はそのまま。
    if isinstance(emp, (int, float)):
        emp = int(emp)
    return {
        "company_name": c.get("企業名", "") or "",
        "industry": c.get("業種_日本語_カテゴリ") or c.get("業種カテゴリ") or "",
        "employee_count": emp,
        "address": c.get("本社住所") or c.get("本社所在地") or "",
    }


def build_page_text(c: dict) -> str:
    """エクセルデータから擬似的なHP本文を構築.

    重要：本シミュレーションは Phase 2 の加点・減点ロジックを検証するためのもの。
    Phase 1 の必須条件（採用情報・健康経営記載）は別レイヤーで判定済みと仮定し、
    全社が必須条件をクリアしているものとして本文を構築する。

    実際の本番では scraper_agent が HP本文を取得して、Phase 1 と Phase 2 を統合判定する。
    """
    # === Phase 1必須条件をクリアする最低限のキーワード（全社共通）===
    # （採用情報・健康経営・福利厚生のいずれか）
    base = (
        "新卒採用・中途採用情報。"          # 採用情報あり
        "健康経営を推進しています。"        # 健康経営記載あり
        "福利厚生として人間ドックを完備。"  # 福利厚生記載あり
    )

    # === Phase 2 シグナル（エクセルデータに応じて変化）===
    parts = [base]

    # 健康経営認定
    if c.get("健康経営認定あり"):
        parts.append("健康経営優良法人2024に認定されています。")

    # 法定外福利厚生（追加分・人間ドックは既に base にある）
    if c.get("法定外福利厚生あり"):
        parts.append("社員旅行、リフレッシュ休暇、マッサージなど充実。")

    # 大手プライム子会社
    if c.get("株主構成タイプ_カテゴリ") == "大手プライム子会社":
        parts.append("東証プライム上場の親会社のもとで事業を展開。")

    # 創業家オーナー
    if c.get("株主構成タイプ_カテゴリ") == "創業家オーナー":
        parts.append("創業家が筆頭株主の同族経営です。")

    # 採用媒体（スコアに応じて媒体名を埋め込む）
    saiyou_score = c.get("採用広告の出稿度_スコア") or 0
    if saiyou_score >= 4:
        parts.append("リクナビ・マイナビ・doda・engageで採用中。")
    elif saiyou_score >= 3:
        parts.append("リクナビ・マイナビ・dodaで採用中。")
    elif saiyou_score >= 2:
        parts.append("リクナビ・マイナビで採用中。")
    elif saiyou_score >= 1:
        parts.append("リクナビで採用中。")

    return " ".join(parts)


def simulate(xlsx_path: str):
    """49社シミュレーション実行."""
    companies = load_companies(xlsx_path)

    keizoku = []
    kaiyaku = []
    others = []

    print(f"分析対象: {len(companies)}社")
    print()
    print(f"{'企業名':<28} {'実際':<14} {'新ﾙｰﾙ':<5} {'点':<4} {'理由(抜粋)':<60}")
    print("=" * 120)

    for c in companies:
        company_info = to_company_info(c)
        page_text = build_page_text(c)

        result = evaluate_rank_v2(
            company_info,
            [{"url": "", "title": "", "snippet": "", "search_query": ""}],
            page_text=page_text,
        )

        actual = c.get("契約状態_カテゴリ") or ""
        rank = result["rank"]
        score = result["score"]
        reasons = ", ".join(result["reasons"][:3])[:58]

        print(f"{c.get('企業名', '')[:27]:<28} {actual[:13]:<14} {rank:<5} {score:>+3}  {reasons}")

        if "継続中" in actual:
            keizoku.append((c.get("企業名", ""), rank, score))
        elif "解約" in actual:
            kaiyaku.append((c.get("企業名", ""), rank, score))
        else:
            others.append((c.get("企業名", ""), rank, score))

    # サマリー
    print()
    print("=" * 60)
    print("【サマリー】")
    print("=" * 60)

    def dist(items):
        d = {"A": 0, "B": 0, "C": 0, "NG": 0}
        for _, r, _ in items:
            d[r] = d.get(r, 0) + 1
        return d

    k_dist = dist(keizoku)
    ka_dist = dist(kaiyaku)
    print(f"\n継続社({len(keizoku)}社) 新ランク分布:")
    for r in ["A", "B", "C", "NG"]:
        cnt = k_dist[r]
        pct = cnt * 100 // len(keizoku) if len(keizoku) > 0 else 0
        print(f"  {r}: {cnt}社 ({pct}%)")

    print(f"\n解約社({len(kaiyaku)}社) 新ランク分布:")
    for r in ["A", "B", "C", "NG"]:
        cnt = ka_dist[r]
        pct = cnt * 100 // len(kaiyaku) if len(kaiyaku) > 0 else 0
        print(f"  {r}: {cnt}社 ({pct}%)")

    # 自動登録精度（3点以上）
    auto_k = sum(1 for _, _, s in keizoku if s >= 3)
    auto_ka = sum(1 for _, _, s in kaiyaku if s >= 3)
    print()
    print("【自動登録（3点以上）の精度】")
    print(f"継続社: {auto_k}/{len(keizoku)}社 ({auto_k*100//max(len(keizoku),1)}%)")
    print(f"解約社: {auto_ka}/{len(kaiyaku)}社 ({auto_ka*100//max(len(kaiyaku),1)}%)")

    # 取りこぼし
    print()
    print("【⚠️ 取りこぼし（継続社なのにNG）】")
    for n, r, s in keizoku:
        if r == "NG":
            print(f"  - {n}: {s}点")

    # 誤検知
    print()
    print("【⚠️ 誤検知（解約社なのにA判定）】")
    for n, r, s in kaiyaku:
        if r == "A":
            print(f"  - {n}: {s}点")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("使い方: python tools/simulate_49_companies.py <xlsxファイルパス>")
        print("例: python tools/simulate_49_companies.py /c/Users/user/Downloads/2026年営業分析シート__3_.xlsx")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    if not os.path.exists(xlsx_path):
        print(f"ERROR: ファイルが見つかりません: {xlsx_path}")
        sys.exit(1)

    simulate(xlsx_path)
